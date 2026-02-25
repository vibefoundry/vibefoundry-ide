"""
Metadata generation for input/output data files
"""

from pathlib import Path
from datetime import datetime
from typing import Optional
import os
import re

# Cache: {filepath: (mtime, row_count, columns, column_info)}
# column_info: {col_name: {"type": str, "date_min": str|None, "date_max": str|None}}
_metadata_cache: dict[str, tuple[float, int, list[str], dict]] = {}

# Keywords that suggest a column contains dates
DATE_COLUMN_KEYWORDS = [
    'date', 'time', 'timestamp', 'created', 'updated', 'modified',
    '_at', '_on', 'datetime', 'day', 'month', 'year', 'dob', 'birth',
    'start', 'end', 'begin', 'finish', 'expire', 'deadline', 'due'
]


def is_date_column_name(col_name: str) -> bool:
    """Check if column name suggests it contains dates."""
    col_lower = col_name.lower()
    return any(keyword in col_lower for keyword in DATE_COLUMN_KEYWORDS)


def try_parse_date(value: str) -> Optional[datetime]:
    """Try to parse a string as a date. Returns datetime or None."""
    if not value or not isinstance(value, str):
        return None

    value = value.strip()
    if not value or len(value) < 6 or len(value) > 30:
        return None

    # Skip obvious non-dates
    if value.isdigit() and len(value) < 8:
        return None

    try:
        from dateutil import parser
        # Try parsing with dateutil (very flexible)
        parsed = parser.parse(value, fuzzy=False)
        # Sanity check: year should be reasonable
        if 1900 <= parsed.year <= 2100:
            return parsed
    except:
        pass

    return None


def detect_date_columns_csv(filepath: Path, columns: list[str], sample_size: int = 5) -> dict:
    """
    Detect date columns in a CSV by sampling values.

    Returns:
        Dict of {col_name: {"type": "date"|"text"|..., "date_min": str, "date_max": str}}
    """
    import csv

    column_info = {col: {"type": "text"} for col in columns}

    try:
        with open(filepath, 'r', encoding='utf-8', errors='ignore') as f:
            reader = csv.DictReader(f)
            samples = {col: [] for col in columns}

            # Collect samples
            for i, row in enumerate(reader):
                if i >= sample_size:
                    break
                for col in columns:
                    if col in row and row[col]:
                        samples[col].append(row[col])

            # Analyze each column
            for col in columns:
                col_samples = samples.get(col, [])
                if not col_samples:
                    continue

                # Check if samples parse as dates
                parsed_dates = []
                for val in col_samples:
                    parsed = try_parse_date(val)
                    if parsed:
                        parsed_dates.append(parsed)

                # If most samples parse as dates, it's a date column
                if len(parsed_dates) >= len(col_samples) * 0.6:
                    column_info[col]["type"] = "date"
                # Also check by column name if we got at least one date
                elif parsed_dates and is_date_column_name(col):
                    column_info[col]["type"] = "date"
    except:
        pass

    return column_info


def detect_date_columns_parquet(filepath: Path) -> dict:
    """
    Detect date columns in a Parquet file using schema and sampling.

    Returns:
        Dict of {col_name: {"type": "date"|"datetime"|"text"|..., "date_min": str, "date_max": str}}
    """
    import polars as pl

    column_info = {}

    try:
        lf = pl.scan_parquet(filepath)
        schema = lf.collect_schema()

        for col_name, dtype in schema.items():
            dtype_str = str(dtype).lower()

            if 'date' in dtype_str or 'datetime' in dtype_str or 'time' in dtype_str:
                column_info[col_name] = {"type": "date"}

                # Get min/max for date columns
                try:
                    stats = lf.select([
                        pl.col(col_name).min().alias("min"),
                        pl.col(col_name).max().alias("max")
                    ]).collect()

                    min_val = stats["min"][0]
                    max_val = stats["max"][0]

                    if min_val is not None:
                        column_info[col_name]["date_min"] = str(min_val)
                    if max_val is not None:
                        column_info[col_name]["date_max"] = str(max_val)
                except:
                    pass

            elif 'int' in dtype_str:
                column_info[col_name] = {"type": "integer"}
            elif 'float' in dtype_str:
                column_info[col_name] = {"type": "float"}
            elif 'bool' in dtype_str:
                column_info[col_name] = {"type": "boolean"}
            else:
                # String column - check if it might be a date by name/sampling
                if is_date_column_name(col_name):
                    # Sample a few values
                    try:
                        sample = lf.select(col_name).head(5).collect()
                        parsed_count = 0
                        for val in sample[col_name].to_list():
                            if val and try_parse_date(str(val)):
                                parsed_count += 1
                        if parsed_count >= 3:
                            column_info[col_name] = {"type": "date"}
                        else:
                            column_info[col_name] = {"type": "text"}
                    except:
                        column_info[col_name] = {"type": "text"}
                else:
                    column_info[col_name] = {"type": "text"}
    except:
        pass

    return column_info


def detect_date_columns_excel(filepath: Path, columns: list[str]) -> dict:
    """Detect date columns in Excel files."""
    column_info = {col: {"type": "text"} for col in columns}

    try:
        import polars as pl
        df = pl.read_excel(filepath, n_rows=10)

        for col in df.columns:
            dtype_str = str(df[col].dtype).lower()
            if 'date' in dtype_str or 'datetime' in dtype_str:
                column_info[col] = {"type": "date"}
            elif 'int' in dtype_str:
                column_info[col] = {"type": "integer"}
            elif 'float' in dtype_str:
                column_info[col] = {"type": "float"}
            elif is_date_column_name(col):
                # Sample values
                for val in df[col].head(5).to_list():
                    if val and try_parse_date(str(val)):
                        column_info[col] = {"type": "date"}
                        break
    except:
        pass

    return column_info


def count_csv_rows_fast(filepath: Path) -> int:
    """Count CSV rows without loading into memory - just count newlines."""
    count = 0
    with open(filepath, 'rb') as f:
        # Read in chunks for memory efficiency
        for chunk in iter(lambda: f.read(1024 * 1024), b''):
            count += chunk.count(b'\n')
    return max(0, count - 1)  # Subtract header row


def get_csv_columns_fast(filepath: Path) -> list[str]:
    """Get CSV column names by reading just the first line."""
    with open(filepath, 'r', encoding='utf-8', errors='ignore') as f:
        header = f.readline().strip()
    # Handle quoted columns
    if ',' in header:
        import csv
        from io import StringIO
        reader = csv.reader(StringIO(header))
        return next(reader, [])
    return header.split(',')


def scan_folder_metadata(folder: Path, title: str) -> str:
    """
    Scan a folder and generate metadata text describing data files.
    Uses caching and fast row counting to avoid loading files into memory.

    Args:
        folder: Path to scan
        title: Title for the metadata section

    Returns:
        Formatted metadata string
    """
    lines = [
        f"{title} Metadata",
        f"Folder: {folder}",
        f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
        "=" * 50,
        ""
    ]

    data_extensions = ['.csv', '.xlsx', '.xls', '.parquet']
    data_files = []
    for ext in data_extensions:
        data_files.extend(folder.glob(f"**/*{ext}"))

    if not data_files:
        lines.append("No data files found.")
        return "\n".join(lines)

    for filepath in sorted(data_files):
        try:
            size_mb = filepath.stat().st_size / (1024 * 1024)
            mtime = filepath.stat().st_mtime
            cache_key = str(filepath)

            # Check cache - use cached values if file hasn't changed
            if cache_key in _metadata_cache:
                cached_mtime, cached_rows, cached_cols, cached_info = _metadata_cache[cache_key]
                if cached_mtime == mtime:
                    row_count = cached_rows
                    columns = cached_cols
                    column_info = cached_info
                else:
                    row_count, columns, column_info = None, None, None
            else:
                row_count, columns, column_info = None, None, None

            # If not cached or stale, read file metadata efficiently
            if row_count is None:
                if filepath.suffix == '.csv':
                    # Fast: count newlines, don't parse
                    row_count = count_csv_rows_fast(filepath)
                    columns = get_csv_columns_fast(filepath)
                    column_info = detect_date_columns_csv(filepath, columns)
                elif filepath.suffix in ['.xlsx', '.xls']:
                    # Excel - use openpyxl for row count, Polars for columns
                    try:
                        from openpyxl import load_workbook
                        wb = load_workbook(filepath, read_only=True)
                        ws = wb.active
                        row_count = ws.max_row - 1  # Subtract header
                        # Get columns from first row
                        columns = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
                        columns = [str(c) if c is not None else '' for c in columns]
                        wb.close()
                    except:
                        # Fallback: use Polars to read Excel
                        import polars as pl
                        df = pl.read_excel(filepath)
                        row_count = len(df)
                        columns = df.columns
                    column_info = detect_date_columns_excel(filepath, columns)
                elif filepath.suffix == '.parquet':
                    # Parquet - Polars can scan metadata without loading
                    import polars as pl
                    try:
                        lf = pl.scan_parquet(filepath)
                        columns = lf.collect_schema().names()
                        row_count = lf.select(pl.len()).collect().item()
                    except:
                        # Fallback: eager load
                        df = pl.read_parquet(filepath)
                        row_count = len(df)
                        columns = df.columns
                    column_info = detect_date_columns_parquet(filepath)
                else:
                    continue

                # Update cache
                _metadata_cache[cache_key] = (mtime, row_count, columns, column_info)

            rel_path = filepath.relative_to(folder)
            lines.append(f"File: {rel_path}")
            lines.append(f"  Absolute Path: {filepath}")
            lines.append(f"  Size: {size_mb:.2f} MB")
            lines.append(f"  Rows: {row_count}")

            # Count date columns
            date_cols = [col for col in columns if column_info.get(col, {}).get("type") == "date"]
            if date_cols:
                lines.append(f"  Date Columns: {len(date_cols)}")

            lines.append(f"  Columns ({len(columns)}):")

            for col in columns:
                info = column_info.get(col, {})
                col_type = info.get("type", "text")

                if col_type == "date":
                    date_min = info.get("date_min", "")
                    date_max = info.get("date_max", "")
                    if date_min and date_max:
                        lines.append(f"    - {col} [DATE: {date_min} to {date_max}]")
                    else:
                        lines.append(f"    - {col} [DATE]")
                elif col_type in ("integer", "float", "boolean"):
                    lines.append(f"    - {col} [{col_type.upper()}]")
                else:
                    lines.append(f"    - {col}")

            lines.append("")

        except Exception as e:
            lines.append(f"File: {filepath.name}")
            lines.append(f"  Error reading: {e}")
            lines.append("")

    return "\n".join(lines)


def generate_metadata(project_folder: Path) -> tuple[Optional[str], Optional[str]]:
    """
    Generate metadata files for input and output folders.

    Args:
        project_folder: Root project folder

    Returns:
        Tuple of (input_metadata, output_metadata) strings, or None if folder doesn't exist
    """
    input_folder = project_folder / "input_folder"
    output_folder = project_folder / "output_folder"
    meta_folder = project_folder / "app_folder" / "meta_data"

    input_meta = None
    output_meta = None

    # Only generate metadata if app_folder structure exists (user clicked Build)
    if not meta_folder.parent.exists():
        return input_meta, output_meta

    # Ensure meta folder exists within app_folder
    meta_folder.mkdir(parents=True, exist_ok=True)

    if input_folder.exists():
        input_meta = scan_folder_metadata(input_folder, "Input Folder")
        (meta_folder / "input_metadata.txt").write_text(input_meta)

    if output_folder.exists():
        output_meta = scan_folder_metadata(output_folder, "Output Folder")
        (meta_folder / "output_metadata.txt").write_text(output_meta)

    return input_meta, output_meta
